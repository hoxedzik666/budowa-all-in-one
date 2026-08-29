"""Import arkusza "Material.xlsx" - kontrola krzyzowa dla danych z PDF.

Arkusze i ich role:
  * Studnie - wymiary i rzedne studni, srednice/rzedne doplywow (D1/RD1, D2/RD2),
              katy (K0/K1/K2).
  * Wpusty  - wpusty deszczowe wraz z kolumna "Odbiornik", czyli JAWNYM grafem
              polaczen, ktorego rysunek profilu nie podaje wprost.
  * Wyloty  - wyloty do rowow i zbiornikow, typ umocnienia (KPED).
  * RURY    - gospodarka materialowa: ile w projekcie, ile dojechalo, WZ, daty.

Zasada: PDF jest zrodlem geometrii. Gdy XLSX podaje inna wartosc, NIE nadpisujemy
- zapisujemy rozbieznosc do raportu importu. Braki uzupelniamy.
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from sqlalchemy import delete, select

from app.extensions import db
from app.models.material import rozbierz_opis
from app.models import (
    Connection,
    ImportRun,
    MaterialItem,
    NetworkObject,
    TypObiektu,
    ZrodloDanych,
)

__all__ = ["importuj_xlsx"]

TOL_ROZBIEZNOSCI_M = 0.02


def _tekst(v, limit: int = 128) -> str | None:
    """Pusta komorka ma zostac None, a nie napisem "None"."""
    if v is None:
        return None
    t = str(v).strip()
    return t[:limit] or None


def _naglowki(ws, szukane: str = "PZ", limit: int = 30) -> tuple[int, dict[str, int]]:
    """Znajdz wiersz naglowka (ten z komorka `szukane`) i mape nazwa -> nr kolumny."""
    for r in range(1, min(limit, ws.max_row or 0) + 1):
        wartosci: dict[str, int] = {}
        for c in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                wartosci.setdefault(v.strip(), c)
        if szukane in wartosci:
            return r, wartosci
    return 0, {}


def importuj_xlsx(sciezka: str | Path) -> ImportRun:
    import openpyxl

    from app.services.importer import (
        RE_KOD_W_TEKSCIE,
        _dodaj_polaczenie,
        _num,
        _sha256,
    )

    widziane_polaczenia: set = set()

    sciezka = Path(sciezka)
    bieg = ImportRun(plik=sciezka.name, sha256=_sha256(sciezka), typ_importu="MATERIAL_XLSX")
    db.session.add(bieg)

    # Import bywa uruchamiany wielokrotnie (po korekcie arkusza, po ponownym
    # imporcie profili). Bez tego kasowania kazdy przebieg dokladal komplet
    # polaczen od nowa - w bazie bylo 2442 wiersze zamiast 771, a karta obiektu
    # pokazywala ten sam doplyw trzy razy.
    db.session.execute(
        delete(Connection).where(Connection.zrodlo == ZrodloDanych.XLSX_MATERIAL)
    )
    db.session.flush()

    wb = openpyxl.load_workbook(sciezka, data_only=True)
    ostrzezenia: list[dict] = []
    staty = {"studnie": 0, "wpusty": 0, "wyloty": 0, "materialy": 0,
             "nowe_obiekty": 0, "polaczenia": 0, "rozbieznosci": 0}

    def znajdz(kod: str) -> NetworkObject | None:
        return db.session.scalar(select(NetworkObject).where(NetworkObject.kod == kod))

    def zapewnij(kod: str, typ: TypObiektu) -> NetworkObject:
        ob = znajdz(kod)
        if ob is None:
            ob = NetworkObject(kod=kod, typ=typ, zrodlo=ZrodloDanych.XLSX_MATERIAL)
            db.session.add(ob)
            db.session.flush()
            staty["nowe_obiekty"] += 1
        return ob

    def porownaj(ob: NetworkObject, pole: str, wartosc, etykieta: str) -> None:
        if wartosc is None:
            return
        obecna = getattr(ob, pole)
        if obecna is None:
            setattr(ob, pole, wartosc)
            return
        if abs(float(obecna) - float(wartosc)) > TOL_ROZBIEZNOSCI_M:
            staty["rozbieznosci"] += 1
            ostrzezenia.append({
                "typ": "ROZBIEZNOSC_PDF_XLSX", "obiekt": ob.kod, "pole": etykieta,
                "pdf": float(obecna), "xlsx": float(wartosc),
                "roznica": round(float(wartosc) - float(obecna), 3),
            })

    def czytaj_kod(ws, r: int, hdr: dict) -> str | None:
        # W kilku wierszach kod trafil do kolumny A zamiast B - sprawdzamy obie.
        for kol in (hdr.get("PZ", 2), 2, 1):
            v = ws.cell(row=r, column=kol).value
            if v and isinstance(v, str) and RE_KOD_W_TEKSCIE.match(v.strip()):
                return v.strip()
        return None

    # ------------------------------------------------------------ Studnie
    if "Studnie" in wb.sheetnames:
        ws = wb["Studnie"]
        r0, hdr = _naglowki(ws)
        for r in range(r0 + 1, (ws.max_row or 0) + 1):
            kod = czytaj_kod(ws, r, hdr)
            if not kod:
                continue
            ob = zapewnij(kod, TypObiektu.STUDNIA)

            def g(nazwa):
                return _num(ws.cell(row=r, column=hdr[nazwa]).value) if nazwa in hdr else None

            porownaj(ob, "rzedna_terenu_proj", g("RTp"), "RTp")
            porownaj(ob, "rzedna_dna_studni", g("Rz.d."), "Rz.d. (dno studni)")
            dn = g("Dn")
            if dn and ob.srednica_studni_mm is None:
                # W arkuszu Dn jest w metrach (1.2 / 1.5 / 2.5), w opisie w mm.
                ob.srednica_studni_mm = int(round(dn * 1000)) if dn < 10 else int(dn)
            if "Opis" in hdr and not ob.opis:
                v = ws.cell(row=r, column=hdr["Opis"]).value
                ob.opis = str(v).strip()[:500] if v else None
            if "Uwagi" in hdr and not ob.uwagi:
                v = ws.cell(row=r, column=hdr["Uwagi"]).value
                ob.uwagi = str(v).strip()[:500] if v else None
            rw = g("Rz. Dna rowu")
            if rw is not None:
                ob.rzedna_dna_rowu = rw

            for kol_dn, kol_rz, kierunek in (
                ("D1", "RD1", "DOPLYW"), ("D2", "RD2", "DOPLYW"),
                ("Dw1", "Rw1", "ODPLYW"), ("Dw2", "Rw2", "ODPLYW"),
            ):
                dn_v, rz_v = g(kol_dn), g(kol_rz)
                if dn_v and rz_v and dn_v >= 50:
                    dodano = _dodaj_polaczenie(Connection(
                        obiekt_id=ob.id, dn_mm=int(dn_v), rzedna=rz_v, kierunek=kierunek,
                        opis=f"{kol_dn}/{kol_rz} wg arkusza Studnie",
                        zrodlo=ZrodloDanych.XLSX_MATERIAL,
                    ), widziane_polaczenia)
                    staty["polaczenia"] += int(dodano)
            staty["studnie"] += 1

    # ------------------------------------------------------------- Wpusty
    if "Wpusty" in wb.sheetnames:
        ws = wb["Wpusty"]
        r0, hdr = _naglowki(ws)
        for r in range(r0 + 1, (ws.max_row or 0) + 1):
            kod = czytaj_kod(ws, r, hdr)
            if not kod:
                continue
            ob = zapewnij(kod, TypObiektu.WPUST)

            def g(nazwa):
                return _num(ws.cell(row=r, column=hdr[nazwa]).value) if nazwa in hdr else None

            porownaj(ob, "rzedna_terenu_proj", g("RTp"), "RTp")
            porownaj(ob, "rzedna_dna_studni", g("Rz.d."), "Rz.d. (dno wpustu)")
            if "Odbiornik" in hdr:
                odb = ws.cell(row=r, column=hdr["Odbiornik"]).value
                if odb and str(odb).strip():
                    dodano = _dodaj_polaczenie(Connection(
                        obiekt_id=ob.id, obiekt_zrodlowy_kod=str(odb).strip()[:64],
                        kierunek="ODPLYW", rzedna=g("Rz"), dn_mm=int(g("D1")) if g("D1") else None,
                        opis="odbiornik wg arkusza Wpusty",
                        zrodlo=ZrodloDanych.XLSX_MATERIAL,
                    ), widziane_polaczenia)
                    staty["polaczenia"] += int(dodano)
            staty["wpusty"] += 1

    # ------------------------------------------------------------- Wyloty
    if "Wyloty" in wb.sheetnames:
        ws = wb["Wyloty"]
        r0, hdr = _naglowki(ws)
        for r in range(r0 + 1, (ws.max_row or 0) + 1):
            kod = czytaj_kod(ws, r, hdr)
            if not kod:
                continue
            ob = zapewnij(kod, TypObiektu.WYLOT)

            def g(nazwa):
                return _num(ws.cell(row=r, column=hdr[nazwa]).value) if nazwa in hdr else None

            porownaj(ob, "rzedna_terenu_proj", g("RTp"), "RTp")
            porownaj(ob, "rzedna_dna_kanalu", g("Rz.d."), "Rz.d. (dno wylotu)")
            dn = g("Dn")
            if dn and ob.dn_mm is None:
                ob.dn_mm = int(round(dn * 1000)) if dn < 10 else int(dn)
            rw = g("Rz. Dna rowu/zbiornika")
            if rw is not None:
                ob.rzedna_dna_rowu = rw
            for kol in ("Uwagi *", "Uwagi"):
                if kol in hdr and not ob.uwagi:
                    v = ws.cell(row=r, column=hdr[kol]).value
                    if v:
                        ob.uwagi = str(v).strip()[:500]
                    break
            staty["wyloty"] += 1

    # --------------------------------------------------------------- RURY
    nazwa_rury = next((n for n in wb.sheetnames if n.strip().upper() == "RURY"), None)
    if nazwa_rury:
        ws = wb[nazwa_rury]
        r0, hdr = _naglowki(ws, szukane="OPIS POZYCJI")
        db.session.execute(delete(MaterialItem))
        for r in range(r0 + 1, (ws.max_row or 0) + 1):
            opis = ws.cell(row=r, column=hdr.get("OPIS POZYCJI", 2)).value
            if not opis or not str(opis).strip():
                continue

            def kom(nazwa, domyslna):
                return ws.cell(row=r, column=hdr.get(nazwa, domyslna)).value

            tekst = str(opis).strip()
            # Srednica, dlugosc sztuki i klasa SN sa w nazwie pozycji - kolumna
            # DLUGOSC bywa pusta, wiec nazwa jest pewniejszym zrodlem.
            rozbior = rozbierz_opis(tekst)
            if rozbior["dlugosc_sztuki_m"] is None:
                rozbior["dlugosc_sztuki_m"] = _num(kom("DŁUGOŚĆ", 4))

            db.session.add(MaterialItem(
                opis_pozycji=tekst[:255],
                dn_od_mm=rozbior["dn_od_mm"],
                dlugosc_sztuki_m=rozbior["dlugosc_sztuki_m"],
                klasa_sn=rozbior["klasa_sn"],
                ilosc_projekt_m=_num(kom("ILOŚCI [M]", 3)),
                ilosc_zamowiona_m=_num(ws.cell(row=r, column=5).value),  # druga kolumna "ILOŚCI [M]"
                ilosc_dostarczona_m=_num(kom("DOJECHAŁO", 6)),
                data_dostawy=_tekst(kom("DATA DOSTAWY", 7)),
                nr_wz=_tekst(kom("WZ", 8)),
                arkusz=nazwa_rury, wiersz_zrodlowy=r,
            ))
            staty["materialy"] += 1
            if rozbior["dn_od_mm"]:
                staty["rury_rozpoznane"] = staty.get("rury_rozpoznane", 0) + 1

    wb.close()

    # Arkusz uzupelnia braki w rzednych, wiec kontrola jakosci musi polecieć
    # jeszcze raz - dopiero teraz komplet danych jest w bazie.
    from app.services.walidacja import sprawdz_dane

    db.session.flush()
    raport = sprawdz_dane(oznacz=True)
    ostrzezenia.extend(p.to_dict() for p in raport.problemy)
    staty.update(raport.statystyki)
    staty["problemy_wg_kategorii"] = raport.wg_kategorii

    bieg.liczba_obiektow = staty["nowe_obiekty"]
    bieg.ostrzezenia = ostrzezenia[:3000]
    bieg.liczba_ostrzezen = len(ostrzezenia)
    bieg.statystyki = staty
    bieg.zakonczono = datetime.now(timezone.utc)
    db.session.commit()
    return bieg
